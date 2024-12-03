import openpyxl
import requests
from bs4 import BeautifulSoup

# URL de la página
url = "https://fitnessworld.com.mx/product-category/aparatos-gym/"

# Realizar la solicitud HTTP
response = requests.get(url)

# Analizar el contenido HTML con BeautifulSoup
soup = BeautifulSoup(response.content, 'html.parser')

# Listas para almacenar los nombres de los productos y los precios
nombres_productos = []
precios = []

# Seleccionar los productos y sus precios
productos = soup.select("h3.product-name")
precios_html = soup.select("span.woocommerce-Price-amount.amount")

# Extraer nombres de productos
for producto in productos:
    try:
        nombre = producto.text.strip()
        nombres_productos.append(nombre)
    except Exception as e:
        print(f"Error al procesar un nombre: {e}")

# Extraer precios
for precio in precios_html:
    try:
        precio_texto = precio.text.strip()
        # Limpiar el precio (quitar '$' y comas, luego convertirlo a número flotante)
        precio_limpio = precio_texto.replace('$', '').replace(',', '').strip()
        precio_numero = float(precio_limpio)
        precios.append(precio_numero)
    except Exception as e:
        print(f"Error al procesar un precio: {e}")

# Cargar el archivo Excel existente
archivo_excel = 'productos_y_precios.xlsx'
wb = openpyxl.load_workbook(archivo_excel)

# Seleccionar la hoja activa
hoja = wb.active

# Buscar la última columna con datos en la primera fila
ultima_columna = hoja.max_column + 1  # Empezamos después de la última columna con datos

# Añadir los nombres de las nuevas columnas en la primera fila
hoja.cell(row=1, column=ultima_columna, value="Aparatos gimnasio")  # Nombre de la columna para los productos
hoja.cell(row=1, column=ultima_columna + 1, value="Precio Aparatos")  # Nombre de la columna para los precios

# Agregar los nuevos datos después de las columnas existentes
for i, (nombre, precio) in enumerate(zip(nombres_productos, precios), start=2):  # Empezar en la fila 2
    hoja.cell(row=i, column=ultima_columna, value=nombre)  # Añadir el nombre del producto
    hoja.cell(row=i, column=ultima_columna + 1, value=precio)  # Añadir el precio como número

# Guardar los cambios en el archivo de Excel
wb.save('productos_y_precios.xlsx')  # Guardar con el mismo nombre
