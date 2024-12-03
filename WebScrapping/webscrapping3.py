import requests
from bs4 import BeautifulSoup
import openpyxl

# Función para extraer productos y precios desde una URL
def extraer_productos_y_precios(url):
    response = requests.get(url)
    soup = BeautifulSoup(response.content, 'html.parser')
    productos = soup.select(".grid-product")
    nombres_productos = []
    precios = []

    for producto in productos:
        try:
            # Nombre del producto
            nombre = producto.select_one('div.grid-product__title.grid-product__title--heading').text.strip()

            # Precio del producto
            precio = producto.select_one('div.grid-product__price').text.strip()

            # Limpiar el precio (quitar '$' y comas)
            precio_limpio = precio.replace('$', '').replace(',', '').strip()

            # Convertir el precio a número (float)
            precio_numero = float(precio_limpio)

            # Agregar los datos a las listas correspondientes
            nombres_productos.append(nombre)
            precios.append(precio_numero)
        except Exception as e:
            print(f"Error al procesar un producto: {e}")

    return nombres_productos, precios

# URLs de las páginas
url1 = "https://suplementosmty.com/collections/proteinas-veganas"
url2 = "https://suplementosmty.com/collections/proteinas?page=3"

# Extraer datos de ambas URLs
nombres_productos_1, precios_1 = extraer_productos_y_precios(url1)
nombres_productos_2, precios_2 = extraer_productos_y_precios(url2)

# Combinar los datos
nombres_productos = nombres_productos_1 + nombres_productos_2
precios = precios_1 + precios_2

# Cargar el archivo Excel existente
ruta_archivo_excel = 'productos_y_precios.xlsx'
wb = openpyxl.load_workbook(ruta_archivo_excel)

# Seleccionar la hoja activa
ws = wb.active

# Determinar si las columnas "Proteinas" y "Precio Proteinas" existen
ultima_columna = ws.max_column
columna_proteinas = ultima_columna + 1  # Nueva columna para "Proteinas"
columna_precio_proteinas = ultima_columna + 2  # Nueva columna para "Precio Proteinas"

# Crear encabezados si las columnas no existen
if not ws.cell(row=1, column=ultima_columna).value == "Proteinas":
    ws.cell(row=1, column=columna_proteinas, value="Proteinas")
    ws.cell(row=1, column=columna_precio_proteinas, value="Precio Proteinas")

# Buscar la última fila ocupada en la columna de "Proteinas"
fila_inicio = 2  # Supongamos que la fila 1 tiene encabezados
while ws.cell(row=fila_inicio, column=columna_proteinas).value:
    fila_inicio += 1

# Escribir los datos de productos y precios en las columnas existentes
for i, (nombre, precio) in enumerate(zip(nombres_productos, precios), start=fila_inicio):
    ws.cell(row=i, column=columna_proteinas, value=nombre)  # Columna de "Proteinas"
    ws.cell(row=i, column=columna_precio_proteinas, value=precio)  # Columna de "Precio Proteinas"

# Guardar el archivo Excel con los nuevos datos
wb.save('productos_y_precios.xlsx')

print("Datos agregados exitosamente al archivo Excel.")
