import requests
from bs4 import BeautifulSoup
import openpyxl

# URL de la página
url = "https://nutrientesvidalabs.com/collections/todos"
response = requests.get(url)

# Crear el objeto BeautifulSoup para analizar el HTML
soup = BeautifulSoup(response.content, "html.parser")

# Extraer los elementos de nombre y precio
productos = soup.select(".product-grid-item")
nombres_productos = []
precios = []

for producto in productos:
    try:
        # Nombre del producto
        nombre = producto.select_one("a.product-grid-item__title").text.strip()

        # Precio del producto
        precio = producto.select_one("a.product-grid-item__price").text.strip()
        
        # Eliminar 'Desde $' del precio y convertirlo en número
        precio = precio.replace('Desde $', '').strip()
        precios.append(float(precio.replace(',', '').strip()))  # Convertir a float

        # Agregar el nombre del producto a la lista
        nombres_productos.append(nombre)
    except Exception as e:
        print(f"Error al procesar un producto: {e}")

# Cargar el archivo de Excel
archivo_excel = 'productos_y_precios.xlsx'  # Nombre del archivo actualizado
wb = openpyxl.load_workbook(archivo_excel)

# Seleccionar la hoja activa
hoja = wb.active

# Buscar la primera columna vacía en la primera fila
ultima_columna = hoja.max_column + 1  # Empezamos después de la última columna con datos

# Añadir los nombres de las nuevas columnas en la primera fila
hoja.cell(row=1, column=ultima_columna, value="nutrientes")  # Nombre de la columna para los productos
hoja.cell(row=1, column=ultima_columna + 1, value="precio nutrientes")  # Nombre de la columna para los precios

# Agregar los nuevos datos después de las columnas existentes
for i, (nombre, precio) in enumerate(zip(nombres_productos, precios), start=2):  # Empezar en la fila 2
    hoja.cell(row=i, column=ultima_columna, value=nombre)  # Añadir el nombre del producto
    hoja.cell(row=i, column=ultima_columna + 1, value=precio)  # Añadir el precio como número

# Guardar los cambios en el archivo de Excel
wb.save('productos_y_precios.xlsx')  # Guardar con un nuevo nombre
