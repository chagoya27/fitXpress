import requests
from bs4 import BeautifulSoup
import openpyxl

# URL de la página
url = "https://www.nike.com/mx/w/entrenamiento-y-gimnasio-prendas-58jtoz6ymx6"

# Encabezado para evitar bloqueos
headers = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/117.0.0.0 Safari/537.36"
}

# Solicitud a la página
response = requests.get(url, headers=headers)
response.raise_for_status()  # Para manejar errores de solicitud
soup = BeautifulSoup(response.text, "html.parser")

# Seleccionar nombres de productos
nombres_productos = [
    nombre.text.strip()
    for nombre in soup.find_all("div", class_="product-card__subtitle")
]

# Seleccionar precios de productos (buscando en ambas etiquetas posibles)
precios_productos = []
for producto in soup.find_all("div", class_="product-card"):
    # Intentar extraer precio de la etiqueta principal
    precio = producto.find("div", class_="product-price is--current-price css-1ydfahe")
    if precio:
        precios_productos.append(precio.text.strip())
    else:
        # Intentar extraer precio de la etiqueta alternativa
        precio_alt = producto.find("div", class_="product-card__price-wrapper")
        precios_productos.append(precio_alt.text.strip() if precio_alt else "N/A")

# Imprimir resultados
print("Nombres de productos:")
print(nombres_productos)

print("\nPrecios de productos:")
print(precios_productos)

# Cargar el archivo Excel existente
archivo = "productos_y_precios.xlsx"
wb = openpyxl.load_workbook(archivo)
hoja = wb.active

# Buscar las columnas para "ropa deportiva" y "precios" (suponiendo que ya existen)
columna_ropa = None
columna_precios = None

# Buscar las columnas de encabezado "ropa deportiva" y "precios"
for col in range(1, hoja.max_column + 1):
    if hoja.cell(row=1, column=col).value == "ropa deportiva":
        columna_ropa = col
    if hoja.cell(row=1, column=col).value == "precios":
        columna_precios = col

# Si no existen las columnas, crearlas
if columna_ropa is None:
    columna_ropa = hoja.max_column + 1
    hoja.cell(row=1, column=columna_ropa, value="ropa deportiva")

if columna_precios is None:
    columna_precios = columna_ropa + 1
    hoja.cell(row=1, column=columna_precios, value="precios")

# Insertar los datos en las columnas correspondientes
for i, (nombre, precio) in enumerate(zip(nombres_productos, precios_productos), start=2):  # Comienza en la fila 2
    hoja.cell(row=i, column=columna_ropa, value=nombre)
    # Convertir precio en formato numérico, eliminando símbolos de moneda y comas
    precio_numerico = precio.replace("$", "").replace(",", "").strip()
    hoja.cell(row=i, column=columna_precios, value=float(precio_numerico) if precio_numerico.replace('.', '', 1).isdigit() else precio)

# Guardar el archivo con los nuevos datos
wb.save(archivo)
print(f"Datos agregados a '{archivo}' exitosamente.")
